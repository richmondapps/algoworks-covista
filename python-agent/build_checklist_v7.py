"""Build R2C_Final_Checklist_v6_apr28.xlsx — Manvitha 4/28 PM feedback applied:

  1. Reframed REQUIREMENTS-first (logic before data points), so traceability runs
     Requirement -> Logic -> Source -> Field -> Validation.
  2. is_accredited gating added to BOTH Course Registration AND Course Participation
     (Bryan's "the accredited piece", Manvitha 4/28).
  3. course_drop explicitly referenced for Course Registration (separate activity_name,
     same activity_log table; invalidates registration if a later drop exists).
  4. "Priority" renamed to "Risk-Tier Weight" with explicit footnote: ALL items are
     validated; weight only controls which incomplete items downgrade the tier
     (per Manvitha's 4/24 rule). Open question to Manvitha: keep weighted or flatten.
  5. Color + live formula scaffolding from v5 retained.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

DST = r"screeshots/apr27/R2C_Final_Checklist_v7_apr28.xlsx"

wb = openpyxl.Workbook()

# =================== SHEET 1: Requirements -> Data Mapping ===================
ws = wb.active
ws.title = "Requirements Mapping"

headers = [
    "#", "Requirement (UI Checklist Item)",
    "Logic / Rule (business intent)",
    "Source Table",
    "Field(s) / Activity Name",
    "Validation Rule (how we prove it)",
    "Risk-Tier Weight*",
    "Status (0/1)",
    "Impacts Readiness?",
]

# Requirements-first: each row reads "the requirement is X, the logic is Y, sourced from Z field"
rows = [
    [
        1, "Initial Portal Login",
        "Student has logged into the student portal at least once",
        "activity_log",
        "activity_name = 'portal_login'",
        "Pending ingestion (Alpesh) - portal logs live in a separate GCP project, no ETA. Treat as 0 until pipeline lands.",
        "HIGH", 0, "YES",
    ],
    [
        2, "FAFSA / Funding Plan Complete",
        "Student has either submitted FAFSA OR is on an alternate funding path",
        "salesforce + activity_log",
        "salesforce.received_fafsa_application_c  |  activity_log.activity_name = 'alternate_funding_submission'",
        "fafsa_submission_flag = TRUE  OR  any 'alternate_funding_submission' row exists for student. Strict boolean; pending/withdrawn -> FALSE (Manvitha 4/27). v17.9 sec 11.",
        "HIGH", 0, "YES",
    ],
    [
        3, "Course Registration (Accredited & Active)",
        "Student is registered for an ACCREDITED course AND has not dropped it",
        "activity_log",
        "activity_name = 'first_course_registration'  +  is_accredited  +  activity_name = 'course_drop' (separate row)",
        "Row exists where activity_name='first_course_registration' AND is_accredited=TRUE, AND NO later activity_name='course_drop' row for the same student/course. Manvitha 4/28: confirm course_drop is its own activity_name (it is - separate row in same table, not a separate column).",
        "HIGH", 0, "YES",
    ],
    [
        4, "No Active Contingencies",
        "No outstanding contingency holds against the student",
        "salesforce",
        "contingency_status",
        "Default TRUE when no active contingency rows (Bryan 4/27: 'no active contingencies' interpretation OK'd). Flip to 0 only when an open contingency row exists.",
        "HIGH", 1, "YES",
    ],
    [
        5, "LMS Login",
        "Student has accessed the LMS at least once (rolling engagement signal)",
        "activity_log",
        "activity_name = 'lms_login'",
        "Any row exists. LIVE in source as of 4/27 (5,423 rows dev + QA, repeats per student per v17.9 sec 13).",
        "HIGH", 0, "YES",
    ],
    [
        6, "WoW Orientation Login",
        "Student has completed Walden's Way orientation login",
        "activity_log",
        "activity_name IN ('wow_login', 'wwow_login')",
        "Either name satisfies (Bryan 4/27: 'leave it, cover both'). 4,082 wwow_login rows currently.",
        "HIGH", 0, "YES",
    ],
    [
        7, "Course Login (Accredited, Post-Start)",
        "Student has logged into an ACCREDITED course AFTER program start date",
        "activity_log",
        "activity_name = 'logged_into_course'  +  is_accredited  +  program_start_date gate",
        "Row exists where activity_name='logged_into_course' AND is_accredited=TRUE AND today >= program_start_date. Suppress pre-start (EVENT_LAYER_SPEC sec 6.5). TRIAGE OPEN: Abraham A01301412 false-positive 28d pre-start (Jake 4/27: raw=TRUE -> source bug to Alpesh; raw=FALSE -> UI bug to Jake).",
        "HIGH", 0, "YES",
    ],
    [
        8, "Course Participation (Post-Start)",
        "Student has submitted to discussion board AFTER program start (program-level, not course-specific)",
        "activity_log",
        "activity_name = 'discussion_board_submission'  +  program_start_date gate",
        "Row exists where activity_name='discussion_board_submission' AND today >= program_start_date. ** NO is_accredited gate ** - Alpesh 4/28: 'discussion_board_submission is not tied to a specific course, rather it is for the entire program.' Field is not populated on these rows by design.",
        "HIGH", 0, "YES",
    ],
]

ws.append(headers)
for r in rows:
    ws.append(r)

last_data_row = ws.max_row  # 9

# Header style
header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True, color="FFFFFF")
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border

# Data row styling
for row in ws.iter_rows(min_row=2, max_row=last_data_row):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border

# Conditional formatting: Risk-Tier Weight col = G; Status col = H
red_fill    = PatternFill("solid", fgColor="F8CBAD")
yellow_fill = PatternFill("solid", fgColor="FFE699")
green_fill  = PatternFill("solid", fgColor="C6EFCE")

weight_range = f"G2:G{last_data_row}"
ws.conditional_formatting.add(weight_range, CellIsRule(operator="equal", formula=['"HIGH"'],   fill=red_fill,    font=Font(bold=True, color="9C0006")))
ws.conditional_formatting.add(weight_range, CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=yellow_fill, font=Font(bold=True, color="9C5700")))
ws.conditional_formatting.add(weight_range, CellIsRule(operator="equal", formula=['"LOW"'],    fill=green_fill,  font=Font(bold=True, color="006100")))

stat_range = f"H2:H{last_data_row}"
ws.conditional_formatting.add(stat_range, CellIsRule(operator="equal", formula=['0'], fill=red_fill))
ws.conditional_formatting.add(stat_range, CellIsRule(operator="equal", formula=['1'], fill=green_fill))

# Column widths
widths = {"A": 4, "B": 32, "C": 42, "D": 22, "E": 50, "F": 70, "G": 13, "H": 11, "I": 13}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Footnote on weight column
ws.append([])
fn_row = ws.max_row + 1
ws.cell(row=fn_row, column=1, value="*").font = Font(bold=True)
ws.cell(row=fn_row, column=2,
    value="ALL items are validated regardless of weight. 'Risk-Tier Weight' only drives which incomplete items downgrade the computed tier (per Manvitha 4/24 rule). Open question: keep weighted scheme or flatten to single tier?"
).font = Font(italic=True, color="595959")
ws.merge_cells(start_row=fn_row, start_column=2, end_row=fn_row, end_column=9)

# Readiness Logic block
ws.append([])
logic_start = ws.max_row + 1
ws.cell(row=logic_start,     column=1, value="Risk-Tier Logic (Manvitha 4/24)").font = Font(bold=True, size=12)
ws.cell(row=logic_start + 1, column=1, value="LOW Risk:").font = Font(bold=True, color="9C0006")
ws.cell(row=logic_start + 1, column=2, value="If ANY HIGH-weight item Status = 0")
ws.cell(row=logic_start + 2, column=1, value="MEDIUM Risk:").font = Font(bold=True, color="9C5700")
ws.cell(row=logic_start + 2, column=2, value="If NO HIGH = 0 but ANY MEDIUM = 0")
ws.cell(row=logic_start + 3, column=1, value="HIGH (Ready):").font = Font(bold=True, color="006100")
ws.cell(row=logic_start + 3, column=2, value="If no HIGH or MEDIUM incomplete")

# Live computed scoring
score_row = logic_start + 5
ws.cell(row=score_row, column=1, value="Live Score").font = Font(bold=True, size=12)

ws.cell(row=score_row + 1, column=1, value="HIGH incomplete count:").font = Font(bold=True)
ws.cell(row=score_row + 1, column=2, value=f'=COUNTIFS(G2:G{last_data_row},"HIGH",H2:H{last_data_row},0)')

ws.cell(row=score_row + 2, column=1, value="MEDIUM incomplete count:").font = Font(bold=True)
ws.cell(row=score_row + 2, column=2, value=f'=COUNTIFS(G2:G{last_data_row},"MEDIUM",H2:H{last_data_row},0)')

ws.cell(row=score_row + 3, column=1, value="Computed Risk Tier:").font = Font(bold=True, size=12)
risk_formula = (
    f'=IF(COUNTIFS(G2:G{last_data_row},"HIGH",H2:H{last_data_row},0)>0,"LOW",'
    f'IF(COUNTIFS(G2:G{last_data_row},"MEDIUM",H2:H{last_data_row},0)>0,"MEDIUM",'
    f'"HIGH"))'
)
risk_cell = ws.cell(row=score_row + 3, column=2, value=risk_formula)
risk_cell.font = Font(bold=True, size=12)
risk_range = f"B{score_row + 3}"
ws.conditional_formatting.add(risk_range, CellIsRule(operator="equal", formula=['"LOW"'],    fill=red_fill,    font=Font(bold=True, color="9C0006")))
ws.conditional_formatting.add(risk_range, CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=yellow_fill, font=Font(bold=True, color="9C5700")))
ws.conditional_formatting.add(risk_range, CellIsRule(operator="equal", formula=['"HIGH"'],   fill=green_fill,  font=Font(bold=True, color="006100")))

ws.freeze_panes = "A2"

# =================== SHEET 2: Business Rules ===================
ws2 = wb.create_sheet("Business Rules")
ws2.append(["Rule", "Description"])
business_rules = [
    ["Accredited Gating",     "Course Registration and Course Login only count when is_accredited=TRUE on the underlying activity_log row (course-level events). Bryan 4/27 'the accredited piece'; Manvitha 4/28 reinforced. NOTE: Course Participation (discussion_board_submission) is program-level NOT course-level per Alpesh 4/28, so is_accredited does not apply there."],
    ["Course Drop Handling",  "course_drop is a separate activity_name (NOT a separate column) in activity_log. If a course_drop row exists AFTER a first_course_registration row for the same student/course, the registration requirement falls back to incomplete until re-registration."],
    ["Temporal Validation",   "Ignore LMS-related events before program_start_date (gates Course Login + Course Participation; EVENT_LAYER_SPEC sec 6.5)."],
    ["Census Rule",           "Remove student if today >= program_start_date + 1 day past census (Bryan 4/27 @ 11:30, supersedes earlier 10-day figure)."],
    ["Negative Time Display", "If today > program_start_date show negative days post-start; zero-day handled separately."],
    ["FAFSA OR Alt-Funding",  "Funding requirement passes if EITHER fafsa_submission_flag=TRUE OR an alternate_funding_submission row exists (v17.9 sec 11)."],
    ["WoW Dual Match",        "Match BOTH wow_login and wwow_login activity names (Bryan 4/27)."],
    ["Email/SMS Draft Blank", "NOT a bug - async AI generation populates 5-7s after page load (Jake 4/27 @ 30:00)."],
    ["Test Environment",      "QA = stable verification env once Jake permissions unblock (Kevin + Vishnu); supersedes 'verify in dev'."],
    ["Course Login Triage",   "If raw logged_into_course=TRUE for pre-start student -> source bug -> Alpesh; if FALSE -> UI bug -> Jake (Jake 4/27 @ 29:19)."],
    ["NBA Grammarly Copy",    "Header: 'Set Up a Free Grammarly Account'; Sub: 'Guide student to Grammarly'; visible link styling."],
    ["Risk-Tier Weight",      "ALL items validated. Weight (HIGH/MEDIUM/LOW) only controls tier downgrade per Manvitha's 4/24 rule. Open: keep weighted or flatten to single tier."],
    ["Discussion Board scope",  "Alpesh confirmed 4/28 PM: discussion_board_submission is PROGRAM-LEVEL (not tied to a specific course). is_accredited is intentionally NOT populated on these rows. Course Participation requirement therefore validates only on row existence + post-start gate."],
]
for r in business_rules:
    ws2.append(r)
for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border
ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 110
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
ws2.freeze_panes = "A2"

# =================== SHEET 3: Change Log ===================
ws3 = wb.create_sheet("Change Log")
ws3.append(["Date", "Author", "Change"])
log = [
    ["2026-04-24", "Manvitha (Teams)",            "Defined risk-tier logic: LOW/MEDIUM/HIGH from HIGH+MEDIUM checklist incompletes"],
    ["2026-04-27", "Nagendra (v2)",              "Initial structure: Checklist Mapping + Business Rules"],
    ["2026-04-28", "Nagendra (v4)",              "Added Priority, Status (0/1), Impacts Readiness columns + readiness logic block"],
    ["2026-04-28", "Nagendra (v5)",               "Color-coded Priority/Status; live formulas wire Status -> Computed Risk Tier"],
    ["2026-04-28", "Manvitha PM feedback",        "Reframe requirements-first; verify is_accredited on Course Reg + Course Participation; reference course_drop field; clarify Priority"],
    ["2026-04-28", "Nagendra (v6)",               "Reframed Requirement -> Logic -> Source -> Field; added is_accredited gating to Course Reg / Login / Participation; added course_drop invalidation rule; renamed Priority -> Risk-Tier Weight w/ footnote"],
    ["2026-04-28", "Alpesh (Slack reply)",        "Clarified discussion_board_submission is PROGRAM-level not course-level; is_accredited not populated on those rows by design"],
    ["2026-04-28", "Nagendra (v7)",               "Removed is_accredited gate from Course Participation (row 8); kept on Course Reg + Course Login only; updated Business Rules + added Discussion Board scope entry"],
]
for r in log:
    ws3.append(r)
for cell in ws3[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 100
for row in ws3.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
ws3.freeze_panes = "A2"

wb.save(DST)
print("Saved:", DST)

# Verify
wb2 = openpyxl.load_workbook(DST)
for s in wb2.worksheets:
    print(f"\n=== SHEET: {s.title} ===")
    for row in s.iter_rows(values_only=True):
        print(" | ".join(str(c) if c is not None else "" for c in row))
