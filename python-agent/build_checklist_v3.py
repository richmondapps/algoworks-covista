"""Build R2C_Checklist_Mapping_v3_apr28.xlsx mirroring ChatGPT v2 structure
(Checklist Mapping + Business Rules sheets) with Apr 27 call corrections applied.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DST = r"screeshots/apr27/R2C_Checklist_Mapping_v3_apr28.xlsx"

wb = openpyxl.Workbook()

# --- Sheet 1: Checklist Mapping ---
ws = wb.active
ws.title = "Checklist Mapping"

headers = ["Checklist Item", "Source Table", "Field / Activity", "Logic", "Validation Rule"]
rows = [
    ["Initial Portal Login",  "activity_log",      "portal_login",                    "exists",                    "Pending ingestion (Alpesh) - portal logs in separate GCP project, no ETA"],
    ["FAFSA Submission",      "salesforce",        "received_fafsa_application_c",    "TRUE OR alt-funding path",  "Strict boolean OR alt-funding (intend_to_fund != 'Federal Aid'); v17.9 sec 11"],
    ["Course Registration",   "activity_log",      "course_registration",             "exists",                    ""],
    ["No Contingencies",      "salesforce",        "contingency_status",              "no active contingencies = TRUE", "Default state TRUE; Bryan 4/27"],
    ["LMS Login",             "activity_log",      "lms_login",                       "exists",                    "Live in source as of 4/27 (5,423 rows dev + QA)"],
    ["WoW Login",             "activity_log",      "wow_login OR wwow_login",         "exists (either)",           "Bryan 4/27: keep dual check, business uses both interchangeably (4,082 rows)"],
    ["Course Login",          "activity_log",      "logged_into_course",              "exists AND today >= program_start_date", "TRIAGE OPEN (Abraham A01301412 28d pre-start FP). Jake 4/27: raw=TRUE -> source bug (Alpesh); raw=FALSE -> UI bug (Jake). Gate per EVENT_LAYER_SPEC sec 6.5"],
    ["Course Participation",  "activity_log",      "discussion_board_submission",     "exists AND today >= program_start_date", "Jake confirmed live in 4/27 call: 'That all looks good.'"],
]

ws.append(headers)
for r in rows:
    ws.append(r)

# Format header
header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True)

# Column widths + wrap text
widths = {"A": 26, "B": 18, "C": 32, "D": 36, "E": 70}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# --- Sheet 2: Business Rules ---
ws2 = wb.create_sheet("Business Rules")
ws2.append(["Rule", "Description"])
business_rules = [
    ["Temporal Validation",   "Ignore LMS events before program_start_date (gates Course Login + Course Participation; EVENT_LAYER_SPEC sec 6.5)"],
    ["Census Rule",           "Remove student if today >= program_start_date + 1 day past census (Bryan corrected 4/27 @ 11:30 from prior 10-day figure)"],
    ["Negative Time Fix",     "If today > program_start_date show negative days (post-start display); zero-day handled separately"],
    ["FAFSA OR Alt-Funding",  "Checklist passes if EITHER fafsa_submission_flag=TRUE OR alt-funding intent recorded (v17.9 sec 11)"],
    ["WoW Dual Match",        "Match BOTH wow_login and wwow_login activity names (Bryan 4/27: 'leave it for now, cover both')"],
    ["Email/SMS Draft Blank", "NOT a bug - async AI generation populates 5-7s after page load (Jake 4/27 @ 30:00)"],
    ["Test Environment",      "QA = stable verification env once Jake permissions unblock (Kevin + Vishnu); supersedes 'verify in dev'"],
    ["Course Login Triage",   "If raw course_login=TRUE -> source bug -> Alpesh; if FALSE -> UI bug -> Jake (Jake 4/27 @ 29:19)"],
    ["NBA Grammarly Copy",    "Header: 'Set Up a Free Grammarly Account'; Sub: 'Guide student to Grammarly'; visible link styling"],
    ["Open: Accredited Piece","Bryan 4/27 @ 37:57 raised 'the accredited piece' - needs Jake sync 4/28"],
]
for r in business_rules:
    ws2.append(r)

for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True)
ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 100
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# --- Sheet 3: Change Log ---
ws3 = wb.create_sheet("Change Log")
ws3.append(["Date", "Author", "Change"])
log = [
    ["2026-04-27", "ChatGPT (Bryan import)", "Initial v2 structure: Checklist Mapping + Business Rules"],
    ["2026-04-28", "Nagendra (post Apr 27 call)", "Added LMS Login row (live 4/27, 5,423 rows)"],
    ["2026-04-28", "Nagendra",                   "WoW Login: explicit dual-match note (Bryan)"],
    ["2026-04-28", "Nagendra",                   "Course Login: added triage protocol + program-start gating (Jake)"],
    ["2026-04-28", "Nagendra",                   "Course Participation: confirmed by Jake live in call"],
    ["2026-04-28", "Nagendra",                   "FAFSA: switched to OR-logic with alt-funding (v17.9 sec 11)"],
    ["2026-04-28", "Nagendra",                   "Census Rule: 10 days -> 1+ day past census (Bryan correction @ 11:30)"],
    ["2026-04-28", "Nagendra",                   "Added rules: Email/SMS draft, Test env, Course Login triage, NBA copy"],
    ["2026-04-28", "Nagendra",                   "Logged open item: 'accredited piece' (Bryan @ 37:57)"],
]
for r in log:
    ws3.append(r)
for cell in ws3[1]:
    cell.fill = header_fill
    cell.font = header_font
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 26
ws3.column_dimensions["C"].width = 90
for row in ws3.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(DST)
print("Saved:", DST)

# Verify
wb2 = openpyxl.load_workbook(DST)
for s in wb2.worksheets:
    print(f"\n=== SHEET: {s.title} ===")
    for row in s.iter_rows(values_only=True):
        print(" | ".join(str(c) if c is not None else "" for c in row))
