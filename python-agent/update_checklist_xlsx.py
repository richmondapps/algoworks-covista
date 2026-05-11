"""Apply Apr 27 call corrections to checklist_mapping spreadsheet -> v2_apr28."""
import openpyxl
import shutil

SRC = r"screeshots/apr27/checklist_mapping_v1 (1).xlsx"
DST = r"screeshots/apr27/checklist_mapping_v2_apr28.xlsx"

shutil.copy(SRC, DST)
wb = openpyxl.load_workbook(DST)
ws = wb["Checklist Mapping"]

for row in ws.iter_rows(min_row=2):
    item = (row[0].value or "").strip()
    if item == "WOW Login":
        row[0].value = "WoW Login"
        row[1].value = "wow_login OR wwow_login"
        row[3].value = "TRUE if either event exists"
        row[4].value = "Bryan 4/27: keep dual check, business uses both interchangeably"
    elif item == "Course Login":
        row[4].value = (
            "TRIAGE OPEN (Abraham/A01301412 28d pre-start false-positive). "
            "Jake 4/27: if raw course_login=TRUE -> source bug (Alpesh); if FALSE -> UI bug (Jake). "
            "Gate by today >= program_start_date per EVENT_LAYER_SPEC sec 6.5."
        )
    elif item == "Course Participation":
        row[4].value = "Jake confirmed 4/27 live: discussion_board_submission is correct."
    elif item.startswith("FAFSA"):
        row[4].value = "Strict boolean (Manvitha 4/27). OR-logic with alt-funding per v17.9 sec 11."
    elif item == "Initial Portal Login":
        row[4].value = "Pending ingestion - portal logs in separate GCP project, no ETA (Alpesh)."
    elif item == "No Contingencies":
        row[4].value = "Bryan 4/27: no active contingencies = TRUE (default state)."

last = ws.max_row + 2
ws.cell(row=last,     column=1, value="--- Updated 2026-04-28 from Apr 27 call (Bryan + Manvitha + Jake + Alpesh + Nagendra) ---")
ws.cell(row=last + 1, column=1, value="Status:")
ws.cell(row=last + 1, column=2, value='Jake reviewed live in call: "That all looks good." Course Participation mapping confirmed.')
ws.cell(row=last + 2, column=1, value="Open items:")
ws.cell(row=last + 2, column=2, value="1) Course Login false-positive triage  2) Initial Portal Login ingestion (Alpesh)")
ws.cell(row=last + 3, column=1, value="Test env:")
ws.cell(row=last + 3, column=2, value="QA = stable verification env once Jake permissions unblock (Kevin+Vishnu).")
ws.cell(row=last + 4, column=1, value="Source counts (Apr 27 sync):")
ws.cell(row=last + 4, column=2, value="lms_login=5423 rows; wwow_login=4082 rows (validated dev + QA, 802,322 total).")

wb.save(DST)
print("Saved:", DST)
print()
for row in ws.iter_rows(values_only=True):
    print(" | ".join(str(c) if c is not None else "" for c in row))
