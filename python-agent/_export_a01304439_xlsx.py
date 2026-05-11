from google.oauth2.credentials import Credentials
from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json

with open(".adtalem_credentials.json") as f:
    info = json.load(f)
creds = Credentials.from_authorized_user_info(info)
client = bigquery.Client(project="qa-wu-agenticai-app-proj", credentials=creds)

sql = """
SELECT activity_name, activity_datetime
FROM `qa-wu-agenticai-app-proj.covista_demo.r2c_student_activity_log`
WHERE student_id = 'A01304439'
ORDER BY activity_datetime DESC
"""
rows = list(client.query(sql).result())

wb = Workbook()
ws = wb.active
ws.title = "A01304439_activity"

headers = ["#", "activity_datetime (UTC)", "activity_name"]
ws.append(headers)
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="305496")
    c.alignment = Alignment(horizontal="center")

for i, r in enumerate(rows, 1):
    dt = r.activity_datetime
    dt_str = "" if dt is None else dt.strftime("%Y-%m-%d %H:%M:%S")
    ws.append([i, dt_str, r.activity_name])

# summary sheet
ws2 = wb.create_sheet("summary")
ws2.append(["activity_name", "count"])
for c in ws2[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="305496")

counts = {}
for r in rows:
    counts[r.activity_name] = counts.get(r.activity_name, 0) + 1
for name, n in sorted(counts.items(), key=lambda x: -x[1]):
    ws2.append([name, n])
ws2.append(["TOTAL", len(rows)])
ws2[ws2.max_row][0].font = Font(bold=True)
ws2[ws2.max_row][1].font = Font(bold=True)

# column widths
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 32
ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 10

out = "A01304439_activity_qa.xlsx"
wb.save(out)
print(f"wrote {out} ({len(rows)} rows)")
