"""Build R2C_Final_Checklist_v8_apr28.xlsx — Apr 28 PM call corrections applied:

Findings from Apr 28 second meeting (Note transcript + Sam-derived clarification):

  1. CONTINGENCY RULE FIX (BIG): Sam's earlier guidance, re-confirmed by Bryan:
     'Verified' status does NOT mean completed - it means the requirement is
     VALIDATED but still exists. Therefore checklist logic must ignore
     contingency_status entirely and key off PRESENCE of contingency_requirement
     rows. ANY row with contingency_requirement IS NOT NULL -> checklist FALSE.
     v7 said 'open contingency rows' - too vague. Simplified xlsx said 'active' -
     also wrong. After-feedback xlsx (in screeshots/ap28) is correct on this.

  2. STATE-vs-EVENT principle codified (Bryan: 'why aren't we looking at
     course_status?' answered correctly on call but not codified in v7).
     activity_log = logic layer; status fields (course_status, etc.) = display only.

  3. Multi-program contingency scope = OPEN QUESTION (Manvitha 4/28).
     For BSN/MSN students, do we filter contingencies by registered program
     or show all? Flagged in Open Questions sheet for Alpesh.

  4. course_drop granularity = OPEN QUESTION. Does the row carry a course
     identifier or is it student-level only? Flagged for Alpesh.

  5. Jake's UI predicate visibility = OPEN QUESTION. Bryan asked Jake to
     reproduce his actual code's logic (existence-only vs sequence-aware).
     Flagged for Jake (Risk #1 from audit).

  6. Bertrand directive: don't ship parallel sheets. The team's existing
     readiness checklist gets an "Implementation Logic" column merged in.
     This v8 file is the source content for that merge - sheet 4 is a
     paste-ready single-column extract.

Carries forward from v7: Risk-Tier weighting, is_accredited gating on
Course Reg + Course Login (not Participation), course_drop invalidation,
post-start temporal gating, Initial Portal Login + Course Participation
rows that the Simplified xlsx dropped.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

DST = r"context/daily/apr29/R2C_Final_Checklist_v8_2_apr28night.xlsx"

wb = openpyxl.Workbook()

# =================== SHEET 1: Requirements -> Data Mapping ===================
ws = wb.active
ws.title = "Requirements Mapping"

headers = [
    "#", "Requirement (UI Checklist Item)",
    "Logic / Rule (business intent)",
    "Source Table",
    "Field(s) / Activity Name",
    "Validation Rule (how we prove it)",
    "Risk-Tier Weight*",
    "Status (0/1)",
    "Impacts Readiness?",
]

rows = [
    [
        1, "Initial Portal Login",
        "Student has logged into the student portal at least once",
        "activity_log",
        "activity_name = 'portal_login'",
        "Pending ingestion (Alpesh) - portal logs live in a separate GCP project, no ETA. Treat as 0 until pipeline lands.",
        "HIGH", 0, "YES",
    ],
    [
        2, "FAFSA / Funding Plan Complete",
        "Student has either submitted FAFSA OR is on an alternate funding path",
        "salesforce + activity_log",
        "fafsa_submission_flag  |  activity_name = 'alternate_funding_submission'",
        "fafsa_submission_flag = TRUE  OR  any 'alternate_funding_submission' row exists for student. Strict boolean; pending/withdrawn -> FALSE (Manvitha 4/27). v17.9 sec 11.",
        "HIGH", 0, "YES",
    ],
    [
        3, "Course Registration (Accredited & Active)",
        "Student is registered for an ACCREDITED course AND has not dropped it",
        "activity_log",
        "activity_name = 'first_course_registration'  +  is_accredited  +  activity_name = 'course_drop' (separate row)",
        "Row exists where activity_name='first_course_registration' AND is_accredited=TRUE, AND NO later activity_name='course_drop' row for the same student/course. Sequence-aware (NOT existence-only). course_drop granularity = OPEN Q for Alpesh (does row carry course identifier?).",
        "HIGH", 0, "YES",
    ],
    [
        4, "No Active Contingencies",
        "Student has no outstanding contingency requirements",
        "activity_log (or salesforce - Alpesh to confirm table)",
        "contingency_requirement",
        "TRUE iff ZERO rows exist where contingency_requirement IS NOT NULL for the student. CRITICAL: ignore contingency_status entirely. Per Sam (re-confirmed Bryan 4/28): 'Verified' = requirement validated but STILL EXISTS, NOT completed. Any row -> contingency present -> FALSE. Multi-program scope = OPEN Q for Alpesh.",
        "HIGH", 0, "YES",
    ],
    [
        5, "LMS Login",
        "Student has accessed the LMS at least once (rolling engagement signal)",
        "activity_log",
        "activity_name = 'lms_login'",
        "Any row exists. LIVE in source as of 4/27 (5,423 rows dev + QA, repeats per student per v17.9 sec 13).",
        "HIGH", 0, "YES",
    ],
    [
        6, "WoW Orientation Login",
        "Student has completed Walden's Way orientation login",
        "activity_log",
        "activity_name IN ('wow_login', 'wwow_login')",
        "Either name satisfies (Bryan 4/27: 'leave it, cover both'). 4,082 wwow_login rows currently.",
        "HIGH", 0, "YES",
    ],
    [
        7, "Course Login (Accredited, Post-Start)",
        "Student has logged into an ACCREDITED course AFTER program start date",
        "activity_log",
        "activity_name = 'logged_into_course'  +  is_accredited  +  program_start_date gate",
        "Row exists where activity_name='logged_into_course' AND is_accredited=TRUE AND today >= program_start_date. Suppress pre-start (EVENT_LAYER_SPEC sec 6.5). TRIAGE OPEN: Abraham A01301412 false-positive 28d pre-start.",
        "HIGH", 0, "YES",
    ],
    [
        8, "Course Participation (Post-Start)",
        "Student has submitted to discussion board AFTER program start (program-level, not course-specific)",
        "activity_log",
        "activity_name = 'discussion_board_submission'  +  program_start_date gate",
        "Row exists where activity_name='discussion_board_submission' AND today >= program_start_date. ** NO is_accredited gate ** - Alpesh 4/28: 'discussion_board_submission is not tied to a specific course, rather it is for the entire program.'",
        "HIGH", 0, "YES",
    ],
]

ws.append(headers)
for r in rows:
    ws.append(r)
last_data_row = ws.max_row

header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True, color="FFFFFF")
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border
for row in ws.iter_rows(min_row=2, max_row=last_data_row):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border

red_fill    = PatternFill("solid", fgColor="F8CBAD")
yellow_fill = PatternFill("solid", fgColor="FFE699")
green_fill  = PatternFill("solid", fgColor="C6EFCE")

weight_range = f"G2:G{last_data_row}"
ws.conditional_formatting.add(weight_range, CellIsRule(operator="equal", formula=['"HIGH"'],   fill=red_fill,    font=Font(bold=True, color="9C0006")))
ws.conditional_formatting.add(weight_range, CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=yellow_fill, font=Font(bold=True, color="9C5700")))
ws.conditional_formatting.add(weight_range, CellIsRule(operator="equal", formula=['"LOW"'],    fill=green_fill,  font=Font(bold=True, color="006100")))
stat_range = f"H2:H{last_data_row}"
ws.conditional_formatting.add(stat_range, CellIsRule(operator="equal", formula=['0'], fill=red_fill))
ws.conditional_formatting.add(stat_range, CellIsRule(operator="equal", formula=['1'], fill=green_fill))

widths = {"A": 4, "B": 32, "C": 42, "D": 22, "E": 50, "F": 70, "G": 13, "H": 11, "I": 13}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.append([])
fn_row = ws.max_row + 1
ws.cell(row=fn_row, column=1, value="*").font = Font(bold=True)
ws.cell(row=fn_row, column=2,
    value="ALL items are validated regardless of weight. 'Risk-Tier Weight' only drives which incomplete items downgrade the computed tier (Manvitha 4/24). Open: keep weighted or flatten to single tier."
).font = Font(italic=True, color="595959")
ws.merge_cells(start_row=fn_row, start_column=2, end_row=fn_row, end_column=9)

ws.append([])
logic_start = ws.max_row + 1
ws.cell(row=logic_start,     column=1, value="Risk-Tier Logic (Manvitha 4/24)").font = Font(bold=True, size=12)
ws.cell(row=logic_start + 1, column=1, value="LOW Risk:").font = Font(bold=True, color="9C0006")
ws.cell(row=logic_start + 1, column=2, value="If ANY HIGH-weight item Status = 0")
ws.cell(row=logic_start + 2, column=1, value="MEDIUM Risk:").font = Font(bold=True, color="9C5700")
ws.cell(row=logic_start + 2, column=2, value="If NO HIGH = 0 but ANY MEDIUM = 0")
ws.cell(row=logic_start + 3, column=1, value="HIGH (Ready):").font = Font(bold=True, color="006100")
ws.cell(row=logic_start + 3, column=2, value="If no HIGH or MEDIUM incomplete")

score_row = logic_start + 5
ws.cell(row=score_row, column=1, value="Live Score").font = Font(bold=True, size=12)
ws.cell(row=score_row + 1, column=1, value="HIGH incomplete count:").font = Font(bold=True)
ws.cell(row=score_row + 1, column=2, value=f'=COUNTIFS(G2:G{last_data_row},"HIGH",H2:H{last_data_row},0)')
ws.cell(row=score_row + 2, column=1, value="MEDIUM incomplete count:").font = Font(bold=True)
ws.cell(row=score_row + 2, column=2, value=f'=COUNTIFS(G2:G{last_data_row},"MEDIUM",H2:H{last_data_row},0)')
ws.cell(row=score_row + 3, column=1, value="Computed Risk Tier:").font = Font(bold=True, size=12)
risk_formula = (
    f'=IF(COUNTIFS(G2:G{last_data_row},"HIGH",H2:H{last_data_row},0)>0,"LOW",'
    f'IF(COUNTIFS(G2:G{last_data_row},"MEDIUM",H2:H{last_data_row},0)>0,"MEDIUM",'
    f'"HIGH"))'
)
risk_cell = ws.cell(row=score_row + 3, column=2, value=risk_formula)
risk_cell.font = Font(bold=True, size=12)
risk_range = f"B{score_row + 3}"
ws.conditional_formatting.add(risk_range, CellIsRule(operator="equal", formula=['"LOW"'],    fill=red_fill,    font=Font(bold=True, color="9C0006")))
ws.conditional_formatting.add(risk_range, CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=yellow_fill, font=Font(bold=True, color="9C5700")))
ws.conditional_formatting.add(risk_range, CellIsRule(operator="equal", formula=['"HIGH"'],   fill=green_fill,  font=Font(bold=True, color="006100")))
ws.freeze_panes = "A2"

# =================== SHEET 2: Business Rules ===================
ws2 = wb.create_sheet("Business Rules")
ws2.append(["Rule", "Description"])
business_rules = [
    ["State vs Event (NEW v8)",  "Activity_log = LOGIC layer (source of truth for all checklist decisions). Status fields like course_status, course_level = CONTEXT layer (display/enrichment only, NEVER drive checklist logic). Bryan 4/28: 'why aren't we looking at course_status?' - because it's a snapshot that loses event ordering. Activity log preserves sequence which checklist rules require."],
    ["Contingency Rule (FIXED v8)", "ANY row where contingency_requirement IS NOT NULL -> contingency exists -> checklist FALSE. contingency_status is NOT used for decisioning. Per Sam (re-confirmed Bryan 4/28): 'Verified' = requirement validated but STILL EXISTS as a contingency, NOT completed. Values: Verified / Submitted / Not Submitted / NULL - only NULL = no contingency."],
    ["Accredited Gating",     "Course Registration and Course Login require is_accredited=TRUE on the activity_log row (course-level events). Course Participation (discussion_board_submission) is PROGRAM-level per Alpesh 4/28, so is_accredited does NOT apply there."],
    ["Course Drop Handling",  "course_drop is a separate activity_name (NOT a column) in activity_log. If a course_drop row exists AFTER a first_course_registration row for the same student/course, registration falls back to incomplete. Sequence-aware NOT existence-only. Granularity (course-level vs student-level) = OPEN Q for Alpesh."],
    ["Temporal Validation",   "Ignore LMS-related events before program_start_date (gates Course Login + Course Participation; EVENT_LAYER_SPEC sec 6.5)."],
    ["Census Rule",           "Remove student if today >= program_start_date + 1 day past census (Bryan 4/27 @ 11:30, supersedes earlier 10-day figure)."],
    ["Negative Time Display", "If today > program_start_date show negative days post-start; zero-day handled separately."],
    ["FAFSA OR Alt-Funding",  "Funding requirement passes if EITHER fafsa_submission_flag=TRUE OR an alternate_funding_submission row exists (v17.9 sec 11)."],
    ["WoW Dual Match",        "Match BOTH wow_login and wwow_login activity names (Bryan 4/27)."],
    ["Email/SMS Draft Blank", "NOT a bug - async AI generation populates 5-7s after page load (Jake 4/27 @ 30:00)."],
    ["Test Environment",      "QA = stable verification env once Jake permissions unblock (Kevin + Vishnu); supersedes 'verify in dev'."],
    ["Course Login Triage",   "If raw logged_into_course=TRUE for pre-start student -> source bug -> Alpesh; if FALSE -> UI bug -> Jake (Jake 4/27 @ 29:19)."],
    ["NBA Grammarly Copy",    "Header: 'Set Up a Free Grammarly Account'; Sub: 'Guide student to Grammarly'; visible link styling."],
    ["Risk-Tier Weight",      "ALL items validated. Weight (HIGH/MEDIUM/LOW) only controls tier downgrade per Manvitha's 4/24 rule. Open: keep weighted or flatten to single tier."],
    ["Discussion Board scope","Alpesh confirmed 4/28 PM: discussion_board_submission is PROGRAM-LEVEL (not tied to a specific course). is_accredited intentionally NOT populated. Course Participation validates only on row existence + post-start gate."],
    ["Single Sheet Directive (NEW v8)", "Bertrand 4/28: stop maintaining parallel sheets. The team's existing readiness checklist gets an 'Implementation Logic' column merged in. v8 sheet 'Paste-Ready Column' is the merge source. v7 / Simplified / After-Feedback xlsx are reference-only going forward."],
]
for r in business_rules:
    ws2.append(r)
for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border
ws2.column_dimensions["A"].width = 34
ws2.column_dimensions["B"].width = 120
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
ws2.freeze_panes = "A2"

# =================== SHEET 3: Open Questions ===================
ws_oq = wb.create_sheet("Open Questions")
ws_oq.append(["#", "Owner", "Question", "Why it matters", "Asked"])
oq_rows = [
    [1, "Sam (CC Mandisa)",
     "Confirm 'Verified' contingency_status = requirement validated but NOT completed (still counts as open contingency).",
     "Locks the v8 contingency rule. Currently inferred from your earlier walkthrough; team needs an explicit re-confirm because the Simplified xlsx and v7 both got it wrong.",
     "Pending"],
    [2, "Jake",
     "Can you paste the exact predicate code your UI/Firestore uses for Course Registration? Specifically: existence-only or sequence-aware on course_drop?",
     "Risk #1 from apr28 audit. v7/v8 spec is sequence-aware ('course_drop AFTER first_course_registration -> invalidate'). If live UI is existence-only, results diverge from spec for any student who registered then dropped.",
     "Pending"],
    [3, "Jake",
     "Sheena Cuevas - dev BQ has first_course_registration for nursing 6003 (accredited) but QA UI shows 'not registered'. Can you share her QA Firestore doc?",
     "Either UI predicate bug or sync edge case despite 802,322-row parity. Blocks Bryan-level confidence in the QA UI.",
     "Pending"],
    [4, "Alpesh",
     "Does the course_drop row carry a course identifier, or is it student-level only?",
     "Determines whether 'no later course_drop for the same student/course' is implementable as written. If student-level only, a drop of any course invalidates registration of every course.",
     "Pending"],
    [5, "Alpesh",
     "Is contingency_requirement a field in activity_log, salesforce table, or both?",
     "v7 said salesforce; the After-Feedback xlsx (4/28) lists activity_log. Need single source of truth for the v8 query.",
     "Pending"],
    [6, "Alpesh",
     "For multi-program students (BSN + MSN) showing contingencies for both programs, do we filter contingencies to the registered program, or show all?",
     "Manvitha 4/28 raised this watching live data. Affects the COUNT(contingency_requirement)=0 predicate (filtered or not).",
     "Pending"],
    [7, "Manvitha",
     "Risk-Tier Weight: keep weighted (HIGH/MEDIUM/LOW with downgrade rules) or flatten to single tier for Bryan-facing review?",
     "Simplified xlsx dropped weights; v7/v8 retain. Decision affects readability vs auditability tradeoff.",
     "Pending"],
    [8, "Bryan / Bertrand",
     "Approve v8 'Implementation Logic' column merge into the team's existing readiness checklist (Bertrand 4/28 directive).",
     "Avoids parallel-sheet confusion that triggered the 4/28 escalation. v8 sheet 4 is the paste-ready source.",
     "Pending"],
]
for r in oq_rows:
    ws_oq.append(r)
for cell in ws_oq[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border
ws_oq.column_dimensions["A"].width = 4
ws_oq.column_dimensions["B"].width = 18
ws_oq.column_dimensions["C"].width = 60
ws_oq.column_dimensions["D"].width = 60
ws_oq.column_dimensions["E"].width = 12
for row in ws_oq.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
ws_oq.freeze_panes = "A2"

# =================== SHEET 4: Spec vs Jake's Actual Code (RECONCILED v8.2 vs apr29 sync-from-bq.ts) ===================
# Reconciliation of v8.1 critique against the updated sync-from-bq.ts Jake shared at 9:57 PM on 4/28.
# Manvitha's 9:02 PM feedback (lms_login removed from checklist; engagement-only) also incorporated.
ws_gap = wb.create_sheet("Spec vs Jake Actual Code")
ws_gap.append(["#", "Requirement", "v8 Spec Says", "Jake's apr29 sync-from-bq.ts Actually Does", "Status", "Severity"])
gap_rows = [
    [1, "Initial Portal Login",
     "activity_log row exists where activity_name='portal_login'",
     "Predicate now hasActivity(['initial_portal_login','lms_login']) (L460). BUT a synthetic 'initial_portal_login' row is still injected per student at reserve_date + random 1-4hrs (L207-210), so the predicate is effectively always TRUE.",
     "PARTIAL FIX - predicate refactored to read activity rows, but synthetic injection makes it tautological. Need to remove or gate the synthetic row before R2C.",
     "HIGH"],
    [2, "FAFSA / Funding",
     "fafsa_submission_flag=TRUE OR activity_name='alternate_funding_submission' row exists (v17.9 sec 11 OR-rule)",
     "hasActivity(['alternate_funding_submission']) OR logs.some(l.fafsa_submission_flag === true/'True'/'true') (L477-486).",
     "FIXED - matches v17.9 sec 11 OR-rule.",
     "OK"],
    [3, "Course Registration (Accredited & Active)",
     "first_course_registration row AND is_accredited=TRUE AND NO later course_drop for same student/course (sequence-aware)",
     "latestReg = first_course_registration with is_accredited filter; latestDrop = course_drop. registrationSatisfied = !!latestReg && (!latestDrop || latestReg.datetime > latestDrop.datetime). (L432-454)",
     "FIXED - sequence-aware with is_accredited gate. Risk #1 from apr28 audit resolved. Open: course_drop is student-level (no per-course key) - confirm with Alpesh.",
     "OK"],
    [4, "No Active Contingencies",
     "Zero rows where contingency_requirement IS NOT NULL for student. Ignore contingency_status.",
     "is_satisfied = !logs.some(l.contingency_status === 'not submitted') (L494-498).",
     "PARTIAL FIX - moved off trf_form_on_file boolean, but still status-based. Sam directive: 'Verified ne completed' means we should not key off contingency_status at all - use presence-of-row count.",
     "HIGH"],
    [5, "LMS Login",
     "activity_log row exists where activity_name='lms_login' (per v17.9 sec 13)",
     "Not a separate checklist item. lms_login folded into Initial Portal Login predicate (L460).",
     "OUT OF SCOPE - Manvitha 4/28 9:02 PM: 'Let us remove lms_login, it is not part of any checklist reference and is needed for the engagement level.' v17.9 sec 13 to be revisited or descoped.",
     "OK"],
    [6, "WoW Orientation Login",
     "activity_log row where activity_name IN ('wow_login','wwow_login')",
     "hasActivity(['wwow_login','wow_login']) (L429).",
     "OK - matches spec exactly.",
     "OK"],
    [7, "Course Login (Accredited, Post-Start)",
     "logged_into_course row AND is_accredited=TRUE AND today >= program_start_date",
     "Own checklist item. Predicate: logs.some(l.activity_name='logged_into_course' && l.is_accredited === true/'True'). (L513-520)",
     "PARTIAL FIX - de-conflated from participation, is_accredited gate added. Still missing post-start temporal gate (Abraham A01301412 case).",
     "HIGH"],
    [8, "Course Participation (Post-Start)",
     "activity_name='discussion_board_submission' row AND today >= program_start_date",
     "Own checklist item. Predicate: logs.some(l.activity_name='discussion_board_submission' && l.is_accredited === true/'True'). (L526-533)",
     "PARTIAL FIX - correct activity now, is_accredited added. Still missing post-start temporal gate.",
     "HIGH"],
    [9, "QA UI read path",
     "QA UI reads same Firestore subcollections written by sync-from-bq.ts",
     "run-qa-sync.ts seeds contingencies as a SEPARATE subcollection keyed by contingency_requirement (L77). sync-from-bq.ts predicate reads activity_log only, not the contingency subcollection.",
     "OPEN - confirm with Jake whether UI reads contingencies subcollection or only activity_log derived flags. Possible second cause behind Sheena Cuevas discrepancy.",
     "HIGH"],
]
for r in gap_rows:
    ws_gap.append(r)
for cell in ws_gap[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border
ws_gap.column_dimensions["A"].width = 4
ws_gap.column_dimensions["B"].width = 30
ws_gap.column_dimensions["C"].width = 50
ws_gap.column_dimensions["D"].width = 65
ws_gap.column_dimensions["E"].width = 55
ws_gap.column_dimensions["F"].width = 12
sev_range = f"F2:F{ws_gap.max_row}"
ws_gap.conditional_formatting.add(sev_range, CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=red_fill,    font=Font(bold=True, color="9C0006")))
ws_gap.conditional_formatting.add(sev_range, CellIsRule(operator="equal", formula=['"HIGH"'],     fill=yellow_fill, font=Font(bold=True, color="9C5700")))
ws_gap.conditional_formatting.add(sev_range, CellIsRule(operator="equal", formula=['"OK"'],       fill=green_fill,  font=Font(bold=True, color="006100")))
for row in ws_gap.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
ws_gap.freeze_panes = "A2"

# =================== SHEET 5: Paste-Ready Column (Bertrand directive) ===================
ws_pr = wb.create_sheet("Paste-Ready Column")
ws_pr.append(["Checklist Item", "Implementation Logic (paste into existing readiness checklist)"])
paste_rows = [
    ["Initial Portal Login",
     "activity_log row exists where activity_name = 'portal_login'. (BLOCKED: pipeline pending Alpesh - separate GCP project. Treat as 0 until ingestion lands.)"],
    ["FAFSA / Funding Plan Complete",
     "fafsa_submission_flag = TRUE  OR  activity_log row exists where activity_name = 'alternate_funding_submission'. (v17.9 sec 11. Pending/withdrawn -> FALSE.)"],
    ["Course Registration (Accredited & Active)",
     "activity_log row exists where activity_name = 'first_course_registration' AND is_accredited = TRUE; AND no later activity_log row where activity_name = 'course_drop' for same student/course. Sequence-aware (not existence-only). Status fields like course_status are NOT used."],
    ["No Active Contingencies",
     "Zero rows exist where contingency_requirement IS NOT NULL for the student. contingency_status is NOT used for decisioning - 'Verified' means requirement validated but STILL EXISTS (per Sam). ANY row -> contingency present -> FALSE."],
    ["LMS Login",
     "activity_log row exists where activity_name = 'lms_login'. (Live in source: 5,423 rows dev + QA, v17.9 sec 13.)"],
    ["WoW Orientation Login",
     "activity_log row exists where activity_name IN ('wow_login', 'wwow_login'). Either name satisfies."],
    ["Course Login (Accredited, Post-Start)",
     "activity_log row exists where activity_name = 'logged_into_course' AND is_accredited = TRUE AND today >= program_start_date. Pre-start rows suppressed (EVENT_LAYER_SPEC sec 6.5)."],
    ["Course Participation (Post-Start)",
     "activity_log row exists where activity_name = 'discussion_board_submission' AND today >= program_start_date. NO is_accredited gate (program-level per Alpesh 4/28)."],
]
for r in paste_rows:
    ws_pr.append(r)
for cell in ws_pr[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border
ws_pr.column_dimensions["A"].width = 38
ws_pr.column_dimensions["B"].width = 130
for row in ws_pr.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
ws_pr.freeze_panes = "A2"

# =================== SHEET 5: Change Log ===================
ws3 = wb.create_sheet("Change Log")
ws3.append(["Date", "Author", "Change"])
log = [
    ["2026-04-24", "Manvitha (Teams)",            "Defined risk-tier logic: LOW/MEDIUM/HIGH from HIGH+MEDIUM checklist incompletes"],
    ["2026-04-27", "Nagendra (v2)",              "Initial structure: Checklist Mapping + Business Rules"],
    ["2026-04-28", "Nagendra (v4)",              "Added Priority, Status (0/1), Impacts Readiness columns + readiness logic block"],
    ["2026-04-28", "Nagendra (v5)",              "Color-coded Priority/Status; live formulas wire Status -> Computed Risk Tier"],
    ["2026-04-28", "Manvitha PM feedback",        "Reframe requirements-first; verify is_accredited; reference course_drop; clarify Priority"],
    ["2026-04-28", "Nagendra (v6)",              "Reframed Requirement -> Logic -> Source -> Field; added is_accredited gating; course_drop invalidation; Priority -> Risk-Tier Weight"],
    ["2026-04-28", "Alpesh (Slack reply)",        "discussion_board_submission is PROGRAM-level not course-level; is_accredited not populated by design"],
    ["2026-04-28", "Nagendra (v7)",              "Removed is_accredited from Course Participation; updated Business Rules + Discussion Board scope entry"],
    ["2026-04-28", "Apr 28 PM call",              "Bryan escalation: contingency rule undefined; Bertrand: stop parallel sheets; Jake to publish UI predicate code; Sam ref: 'Verified' != completed"],
    ["2026-04-28", "Nagendra (v8)",              "FIXED contingency rule (presence-of-row, not status); added State-vs-Event business rule; added Open Questions sheet; added Paste-Ready Column sheet for Bertrand single-sheet directive"],
    ["2026-04-28", "Nagendra (v8.1 - code audit)", "Read functions/src/sync-from-bq.ts (older repo copy). Flagged 7 CRITICAL divergences from v8 spec."],
    ["2026-04-28", "Nagendra (v8.2 - reconciled)",  "Reconciled v8.1 critique against Jake's apr29 sync-from-bq.ts (attached 9:57 PM, not yet pushed). FAFSA OR-rule FIXED. Course Reg sequence+accredited FIXED. Course Participation activity FIXED. WoW already correct. Portal Login PARTIAL (synthetic row still injected). Contingencies PARTIAL (still status-based, not row-count). Course Login PARTIAL (no post-start gate). LMS Login OUT OF SCOPE per Manvitha 9:02 PM. QA dual-write path OPEN. Net: 4 OK, 4 HIGH partial/open, 0 CRITICAL."],
]
for r in log:
    ws3.append(r)
for cell in ws3[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 110
for row in ws3.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
ws3.freeze_panes = "A2"

wb.save(DST)
print("Saved:", DST)

wb2 = openpyxl.load_workbook(DST)
for s in wb2.worksheets:
    print(f"\n=== SHEET: {s.title} ===")
    for row in s.iter_rows(values_only=True):
        print(" | ".join(str(c) if c is not None else "" for c in row))
