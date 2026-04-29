from openpyxl import load_workbook
wb = load_workbook(r'context/daily/apr29/R2C_Checklist_Split_Logic_included.xlsx')
for s in wb.worksheets:
    print('==SHEET==', s.title)
    for r in s.iter_rows(values_only=True):
        print(r)
