"""Generate simple R2C Checklist (Split Logic style) reconciled with Jake's apr29 sync-from-bq.ts.

Mirrors the format Manvitha liked from ChatGPT's R2C_Checklist_Split_Logic_included.xlsx,
but content is corrected against actual code + Manvitha's lms_login removal.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"context/daily/apr29/R2C_Checklist_Simple_apr29.xlsx"

HEADER = ["#", "Checklist Item", "Type (Blocking/Info)", "Source",
          "Fields / Activity", "Logic", "Status Rule", "Notes"]

ROWS = [
    (1, "Initial Portal Login", "Blocking", "activity_log",
     "activity_name IN ('initial_portal_login','lms_login')",
     "EXISTS portal_login OR lms_login",
     "TRUE if exists",
     "Synthetic initial_portal_login row currently injected per student at reserve_date+1–4hrs (placeholder until real ingestion)."),

    (2, "FAFSA / Funding Plan", "Blocking", "salesforce + activity_log",
     "fafsa_submission_flag OR alternate_funding_submission",
     "fafsa_submission_flag = TRUE  OR  EXISTS alternate_funding_submission",
     "TRUE if either",
     "v17.9 §11 OR-rule. Locked in PR #7."),

    (3, "Course Registration", "Blocking", "activity_log",
     "first_course_registration (is_accredited=TRUE), course_drop",
     "latestReg.datetime > latestDrop.datetime  (sequence-aware, accredited only)",
     "TRUE if latest reg after latest drop",
     "Risk #1 resolved in Jake's apr29 sync-from-bq.ts."),

    (4, "No Active Contingencies", "Blocking", "activity_log",
     "contingency_requirement / contingency_status",
     "Current code: NOT EXISTS row with contingency_status='not submitted'",
     "TRUE if no 'not submitted' rows",
     "Sam directive: should key off COUNT(contingency rows)=0; Verified ≠ completed. OPEN."),

    (5, "Contingencies Identified", "Informational", "activity_log",
     "contingency_requirement, contingency_status",
     "Display all rows",
     "Always shown",
     "Informational only — does NOT block readiness."),

    (6, "WOW Login", "Blocking", "activity_log",
     "activity_name IN ('wow_login','wwow_login')",
     "EXISTS wow login",
     "TRUE if exists",
     None),

    (7, "Course Login", "Blocking", "activity_log",
     "logged_into_course (is_accredited=TRUE), program_start_date",
     "EXISTS logged_into_course AND login.datetime >= program_start_date",
     "TRUE if valid login after start",
     "Post-start gate NOT in current code — OPEN. Abraham A01301412 case."),

    (8, "Course Participation", "Blocking", "activity_log",
     "discussion_board_submission (is_accredited=TRUE), program_start_date",
     "EXISTS discussion_board_submission AND submission.datetime >= program_start_date",
     "TRUE if valid submission after start",
     "Activity corrected to discussion_board_submission. Post-start gate OPEN."),
]

# NOTE: lms_login removed as a separate checklist item per Manvitha (engagement only,
# folded into Initial Portal Login predicate). 8 items total.

wb = Workbook()
ws = wb.active
ws.title = "Checklist Split Logic"

header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True, color="FFFFFF", size=11)
zebra = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(HEADER)
for c in ws[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

for i, row in enumerate(ROWS, start=2):
    ws.append(row)
    for c in ws[i]:
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        if i % 2 == 0:
            c.fill = zebra

widths = [5, 26, 18, 24, 38, 48, 28, 54]
for idx, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + idx)].width = w

ws.row_dimensions[1].height = 32
for r in range(2, len(ROWS) + 2):
    ws.row_dimensions[r].height = 48

ws.freeze_panes = "A2"

wb.save(OUT)
print("WROTE:", OUT)
