"""R2C Checklist — Merged (Apr 29, 7am) update after Data Validation Meeting.

Reflects locked decisions from the Apr 29 7am Data Validation Meeting:

  1. Existence-based logic ONLY for Course Login + Course Participation.
     activity_datetime is frequently NULL in source for student-event rows;
     timestamp gates produce false negatives. Per business SME walkthrough:
     "If a student hasn't performed the activity, no row is emitted —
      so existence of the row is sufficient."
  2. Initial Portal Login: REMOVE lms_login from logic. Portal != LMS.
     LMS login is engagement-only (separate signal), not a checklist item.
     initial_portal_login is currently a synthetic placeholder until real
     ingestion lands.
  3. Course Registration: simplify validation rule to existence + accredited.
     Drop the "latestReg > latestDrop" wording from the rule itself; keep
     course_drop as an OBSERVATION in Notes (priority signal, not an
     invalidator of completion).
  4. Contingencies: rule is OPEN pending business confirmation. Current
     code negates contingency_status='not submitted'; the source may
     pre-filter so any row present = outstanding. Awaiting walkthrough
     with business SME (Sam) and any SF-side checkbox semantics review.
  5. NBA pre-program-start suppression is a UI/surface rule, NOT a
     completion invalidator. Do not block readiness on temporal gates.
  6. Single source of truth = Bertrand's master Readiness Checklist
     (cols K/L/M = Data Satisfaction Rule / Notes / Data Discrepancies).
     This file remains an audit/working artifact; final logic merges
     into Bertrand's master sheet.

Standing rules (carry-over):
  - No name attribution and no tool/source attribution in any shareable
    artifact. "Per business SME walkthrough" is the safe phrasing.

Columns: Requirement | Source Table | Field(s) | Expected Value (Type)
         | Logic (Business) | Validation Rule (Simple) | Status
         | Notes | Data Discrepancies
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"context/daily/apr29/R2C_Checklist_Merged_apr29_7am.xlsx"

HEADER = [
    "Requirement", "Source Table", "Field(s)", "Expected Value (Type)",
    "Logic (Business)", "Validation Rule (Simple)", "Status",
    "Notes", "Data Discrepancies",
]

ROWS = [
    (
        "Initial Portal Login",
        "activity_log",
        "activity_name",
        "STRING",
        "Student logged into the portal (Anthology).",
        "EXISTS activity_name = 'initial_portal_login'",
        "OPEN — placeholder data",
        "LMS login removed from this rule (portal != LMS; LMS is engagement-only). "
        "Synthetic initial_portal_login row currently injected per student until real "
        "ingestion lands.",
        "Real ingestion source pending.",
    ),
    (
        "FAFSA / Funding Plan Complete",
        "salesforce + activity_log",
        "fafsa_submission_flag, activity_name",
        "BOOLEAN, STRING",
        "Student submitted FAFSA OR alternate funding.",
        "fafsa_submission_flag = TRUE  OR  EXISTS activity_name = 'alternate_funding_submission'",
        "OK (locked v17.9 §11)",
        "OR-rule, not AND. Mixed-path students (alt funding + FAFSA) are valid.",
        "None.",
    ),
    (
        "Course Registration",
        "activity_log",
        "activity_name, is_accredited",
        "STRING, BOOLEAN",
        "Student is registered for at least one accredited course.",
        "EXISTS activity_name = 'first_course_registration' AND is_accredited = TRUE",
        "OK (simplified)",
        "course_drop is tracked as an OBSERVATION (priority signal) but does not "
        "invalidate completion at the checklist layer. Sequence-aware downstream handling "
        "(latestReg vs latestDrop) is retained in the sync layer for reporting.",
        "Some students show no first_course_registration; confirmed correct behavior "
        "(event-driven, not behavioral inference). Coverage % to be quantified separately.",
    ),
    (
        "No Active Contingencies",
        "activity_log",
        "contingency_requirement, contingency_status",
        "STRING, STRING",
        "Student has no outstanding contingencies.",
        "PENDING BUSINESS CONFIRMATION on final rule "
        "(status-based vs source-filtered outstanding-only):\n"
        "  (a) NOT EXISTS row with contingency_status = 'not submitted' (current code)\n"
        "  (b) COUNT(rows where contingency_requirement IS NOT NULL) = 0 "
        "(if source pre-filters to outstanding-only)",
        "OPEN — awaiting business confirmation",
        "Per business SME walkthrough: any contingency row surfaced from source should "
        "be treated as outstanding/needed; SF-side may use a checkbox that maps to "
        "'verified'. Display behavior: list each outstanding contingency individually "
        "(e.g., Transcript, Nursing License) per university; show 'No contingencies' "
        "only when none present.",
        "Need to confirm whether 'verified' = satisfied, and whether source already "
        "filters to outstanding rows only. Walkthrough with business SME pending.",
    ),
    (
        "WOW Login",
        "activity_log",
        "activity_name",
        "STRING",
        "Student completed Week of Welcome login.",
        "EXISTS activity_name IN ('wow_login', 'wwow_login')",
        "OK",
        "Both spellings supported (source emits either).",
        "None.",
    ),
    (
        "Course Login",
        "activity_log",
        "activity_name, is_accredited",
        "STRING, BOOLEAN",
        "Student logged into an accredited course.",
        "EXISTS activity_name = 'logged_into_course' AND is_accredited = TRUE",
        "OK (existence-based)",
        "Datetime gate REMOVED from validation rule. Absence of row indicates activity "
        "not performed (no reliance on datetime). NBA pre-start suppression is strictly "
        "a UI-layer rule (does not impact backend completion logic) — no surfacing "
        "earlier than ~3 days before program_start_date.",
        "Activity model: missing activity = no row (not null timestamp). All validation "
        "standardized to EXISTS-based checks.",
    ),
    (
        "Course Participation",
        "activity_log",
        "activity_name, is_accredited",
        "STRING, BOOLEAN",
        "Student submitted a discussion board post.",
        "EXISTS activity_name = 'discussion_board_submission' AND is_accredited = TRUE",
        "OK (existence-based)",
        "Datetime gate REMOVED from validation rule — same rationale as Course Login. "
        "Discussion board submission is the program-level participation signal "
        "(not per-course). NBA pre-start suppression is strictly a UI-layer rule "
        "(does not impact backend completion logic).",
        "Two students reviewed had no row — confirmed correct behavior "
        "(missing activity = no row, not null timestamp).",
    ),
]

wb = Workbook()
ws = wb.active
ws.title = "Readiness Checklist Mapping"

header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True, color="FFFFFF", size=11)
zebra = PatternFill("solid", fgColor="F2F2F2")
ok_fill = PatternFill("solid", fgColor="C6EFCE")
open_fill = PatternFill("solid", fgColor="FFEB9C")
red_fill = PatternFill("solid", fgColor="F8CBAD")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(HEADER)
for c in ws[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

for i, row in enumerate(ROWS, start=2):
    ws.append(row)
    status = (row[6] or "").lower()
    for c in ws[i]:
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        if i % 2 == 0:
            c.fill = zebra
    status_cell = ws.cell(row=i, column=7)
    if status.startswith("ok"):
        status_cell.fill = ok_fill
        status_cell.font = Font(bold=True, color="006100")
    elif "awaiting business" in status or "placeholder" in status:
        status_cell.fill = red_fill
        status_cell.font = Font(bold=True, color="9C0006")
    elif status.startswith("open"):
        status_cell.fill = open_fill
        status_cell.font = Font(bold=True, color="9C5700")

widths = [28, 22, 28, 22, 38, 50, 24, 50, 38]
for idx, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + idx)].width = w

ws.row_dimensions[1].height = 34
for r in range(2, len(ROWS) + 2):
    ws.row_dimensions[r].height = 90

ws.freeze_panes = "A2"

# --- Sheet 2: Change Log -----------------------------------------------------
ws2 = wb.create_sheet("Change Log")
ws2.append(["Version", "Date", "Change"])
for c in ws2[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

changes = [
    ("apr29 (initial merged)", "2026-04-28 night",
     "8-col merged sheet; sequence-aware course reg; datetime gates on Course Login + "
     "Course Participation; lms_login still in Initial Portal Login predicate."),
    ("apr29 7am (this file)", "2026-04-29 morning",
     "Activity model clarification: missing activity = no row (not null timestamp); "
     "all validation standardized to EXISTS-based checks. "
     "Removed datetime gates from Course Login + Course Participation. "
     "Removed lms_login from Initial Portal Login predicate. Simplified Course Registration "
     "rule (existence + is_accredited); course_drop moved to Notes as observation. "
     "Contingencies marked OPEN pending business confirmation on final rule "
     "(status-based vs source-filtered outstanding-only). NBA pre-start suppression "
     "clarified as strictly UI-layer (does not impact backend completion logic). "
     "Added Data Discrepancies column to mirror master sheet structure."),
]
for r in changes:
    ws2.append(r)
    for c in ws2[ws2.max_row]:
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border

ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 110
for r in range(2, ws2.max_row + 1):
    ws2.row_dimensions[r].height = 75

# --- Sheet 3: Open Items -----------------------------------------------------
ws3 = wb.create_sheet("Open Items")
ws3.append(["#", "Item", "Owner", "Status"])
for c in ws3[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

opens = [
    (1, "Contingencies rule: confirm whether source pre-filters to outstanding-only, "
        "and whether 'verified' status = satisfied. Decide between status-negation vs "
        "row-count rule.",
     "Business SME + UI lead", "OPEN"),
    (2, "Initial Portal Login: replace synthetic placeholder with real ingestion when "
        "available.",
     "Data + UI", "OPEN"),
    (3, "NBA pre-program-start suppression window: confirm cutoff "
        "(weekend before / 3 days before).",
     "Product", "OPEN"),
    (4, "Course Registration coverage: quantify % of students missing "
        "first_course_registration to distinguish expected from data-gap cases.",
     "Data", "OPEN"),
    (5, "QA application access: intermittent failure observed; service account / "
        "production read access setup tracked separately.",
     "DevOps", "OPEN"),
]
for r in opens:
    ws3.append(r)
    for c in ws3[ws3.max_row]:
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border

ws3.column_dimensions["A"].width = 6
ws3.column_dimensions["B"].width = 90
ws3.column_dimensions["C"].width = 22
ws3.column_dimensions["D"].width = 12
for r in range(2, ws3.max_row + 1):
    ws3.row_dimensions[r].height = 60

wb.save(OUT)
print("WROTE:", OUT)
