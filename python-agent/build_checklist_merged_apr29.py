"""Merged R2C Checklist (apr29) — ChatGPT column structure + our corrected logic.

Columns: Requirement | Source Table | Field(s) | Expected Value (Type) | Logic (Business)
         | Validation Rule (Simple) | Status | Notes
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"context/daily/apr29/R2C_Checklist_Merged_apr29.xlsx"

HEADER = ["Requirement", "Source Table", "Field(s)", "Expected Value (Type)",
          "Logic (Business)", "Validation Rule (Simple)", "Status", "Notes"]

ROWS = [
    ("Initial Portal Login", "activity_log",
     "activity_name",
     "STRING",
     "Student logged into portal (Anthology) or LMS",
     "activity_name IN ('initial_portal_login','lms_login')",
     "OK (code aligned)",
     "Synthetic initial_portal_login row currently injected per student at reserve_date+1–4hrs — placeholder until real ingestion lands."),

    ("FAFSA / Funding Plan Complete", "salesforce + activity_log",
     "fafsa_submission_flag, activity_name",
     "BOOLEAN, STRING",
     "Student submitted FAFSA OR alternate funding",
     "fafsa_submission_flag = TRUE  OR  EXISTS activity_name = 'alternate_funding_submission'",
     "OK (locked v17.9 §11, PR #7)",
     "OR-rule, not AND."),

    ("Course Registration", "activity_log",
     "activity_name, is_accredited, course_drop datetime",
     "STRING, BOOLEAN, TIMESTAMP",
     "Student is registered for an accredited course AND latest registration is after latest drop",
     "latestReg(first_course_registration, is_accredited=TRUE).datetime > latestDrop.datetime",
     "OK (Risk #1 resolved)",
     "Sequence-aware in Jake's apr29 sync-from-bq.ts."),

    ("No Active Contingencies", "activity_log",
     "contingency_requirement, contingency_status",
     "STRING, STRING",
     "Student has no outstanding contingencies",
     "Current code: NOT EXISTS row with contingency_status = 'not submitted'\n"
     "Business directive: COUNT(contingency_requirement rows) = 0",
     "OPEN — code vs business directive mismatch",
     "Per business SME walkthrough: Verified ≠ completed. Need decision: status-based (current) vs row-count-based (directive)."),

    ("Contingencies Identified (Informational)", "activity_log",
     "contingency_requirement, contingency_status",
     "STRING, STRING",
     "Display all contingency rows for visibility — does NOT block readiness",
     "Display all rows where contingency_requirement IS NOT NULL",
     "Informational",
     "Always shown. Separate from blocking 'No Active Contingencies' rule."),

    ("WOW Login", "activity_log",
     "activity_name",
     "STRING",
     "Student completed Week of Welcome login",
     "activity_name IN ('wow_login','wwow_login')",
     "OK",
     "Both spellings supported in Jake's hasActivity helper."),

    ("Course Login", "activity_log",
     "activity_name, is_accredited, program_start_date",
     "STRING, BOOLEAN, DATE",
     "Student logged into accredited course on/after program start",
     "EXISTS activity_name='logged_into_course' AND is_accredited=TRUE\n"
     "AND login.datetime >= program_start_date",
     "OPEN — post-start gate missing in code",
     "Abraham A01301412 case: pre-start logins should not satisfy. Add temporal gate."),

    ("Course Participation", "activity_log",
     "activity_name, is_accredited, program_start_date",
     "STRING, BOOLEAN, DATE",
     "Student submitted discussion board post on/after program start",
     "EXISTS activity_name='discussion_board_submission' AND is_accredited=TRUE\n"
     "AND submission.datetime >= program_start_date",
     "OPEN — post-start gate missing in code",
     "Activity corrected to discussion_board_submission (was first_assignment_submitted_at in older drafts)."),
]

wb = Workbook()
ws = wb.active
ws.title = "Readiness Checklist Mapping"

header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True, color="FFFFFF", size=11)
zebra = PatternFill("solid", fgColor="F2F2F2")
ok_fill = PatternFill("solid", fgColor="C6EFCE")
open_fill = PatternFill("solid", fgColor="FFEB9C")
info_fill = PatternFill("solid", fgColor="DDEBF7")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(HEADER)
for c in ws[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

for i, row in enumerate(ROWS, start=2):
    ws.append(row)
    status = (row[6] or "").lower()
    for c in ws[i]:
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        if i % 2 == 0:
            c.fill = zebra
    # Status column color override (column G = index 7)
    status_cell = ws.cell(row=i, column=7)
    if status.startswith("ok"):
        status_cell.fill = ok_fill
        status_cell.font = Font(bold=True, color="006100")
    elif status.startswith("open"):
        status_cell.fill = open_fill
        status_cell.font = Font(bold=True, color="9C5700")
    elif status.startswith("informational"):
        status_cell.fill = info_fill
        status_cell.font = Font(bold=True, color="1F4E78")

widths = [30, 22, 30, 24, 42, 50, 28, 50]
for idx, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + idx)].width = w

ws.row_dimensions[1].height = 34
for r in range(2, len(ROWS) + 2):
    ws.row_dimensions[r].height = 60

ws.freeze_panes = "A2"

wb.save(OUT)
print("WROTE:", OUT)
