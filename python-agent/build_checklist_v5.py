"""Build R2C_Final_Checklist_v5_apr28.xlsx — extends ChatGPT v4 with:
  - Apr 27 call corrections (LMS Login row, WoW dual, Course Login triage, FAFSA OR)
  - Color-coded Priority cells (HIGH=red, MEDIUM=yellow, LOW=green)
  - Live scoring formulas (Readiness driven by Status column)
  - Manvitha's Friday risk-tier logic wired to checklist counts
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

DST = r"screeshots/apr27/R2C_Final_Checklist_v5_apr28.xlsx"

wb = openpyxl.Workbook()

# =================== SHEET 1: Checklist Mapping ===================
ws = wb.active
ws.title = "Checklist Mapping"

headers = [
    "Checklist Item", "Source Table", "Field / Activity",
    "Logic", "Validation Rule", "Priority", "Status (0/1)", "Impacts Readiness?"
]
rows = [
    # item, src, field, logic, validation, priority, status, impacts
    ["Initial Portal Login",  "activity_log", "portal_login",                 "exists",                                   "Pending ingestion (Alpesh) - portal logs in separate GCP project, no ETA",          "HIGH",   0, "YES"],
    ["FAFSA Submission",      "salesforce",   "received_fafsa_application_c", "TRUE OR alt-funding path",                 "Strict boolean OR alt-funding (intend_to_fund != 'Federal Aid'); v17.9 sec 11",      "HIGH",   0, "YES"],
    ["Course Registration",   "activity_log", "course_registration",          "exists",                                   "",                                                                                   "HIGH",   0, "YES"],
    ["No Contingencies",      "salesforce",   "contingency_status",           "no active contingencies = TRUE",           "Default state TRUE; Bryan 4/27",                                                     "HIGH",   1, "YES"],
    ["LMS Login",             "activity_log", "lms_login",                    "exists",                                   "Live in source as of 4/27 (5,423 rows dev + QA)",                                    "HIGH",   0, "YES"],
    ["WoW Login",             "activity_log", "wow_login OR wwow_login",      "exists (either)",                          "Bryan 4/27: keep dual check, business uses both interchangeably (4,082 rows)",       "HIGH",   0, "YES"],
    ["Course Login",          "activity_log", "logged_into_course",           "exists AND today >= program_start_date",   "TRIAGE OPEN (Abraham A01301412 28d pre-start FP). Jake 4/27: raw=TRUE -> source bug (Alpesh); raw=FALSE -> UI bug (Jake). Gate per EVENT_LAYER_SPEC sec 6.5", "MEDIUM", 0, "YES"],
    ["Course Participation",  "activity_log", "discussion_board_submission",  "exists AND today >= program_start_date",   "Jake confirmed live in 4/27 call: 'That all looks good.'",                           "MEDIUM", 0, "YES"],
]

ws.append(headers)
for r in rows:
    ws.append(r)

last_data_row = ws.max_row  # row 9

# Header style
header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True, color="FFFFFF")
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border

# Data row styling + borders
for row in ws.iter_rows(min_row=2, max_row=last_data_row):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border

# Conditional formatting on Priority column (F)
red_fill    = PatternFill("solid", fgColor="F8CBAD")  # soft red
yellow_fill = PatternFill("solid", fgColor="FFE699")
green_fill  = PatternFill("solid", fgColor="C6EFCE")
prio_range = f"F2:F{last_data_row}"
ws.conditional_formatting.add(prio_range, CellIsRule(operator="equal", formula=['"HIGH"'],   fill=red_fill,    font=Font(bold=True, color="9C0006")))
ws.conditional_formatting.add(prio_range, CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=yellow_fill, font=Font(bold=True, color="9C5700")))
ws.conditional_formatting.add(prio_range, CellIsRule(operator="equal", formula=['"LOW"'],    fill=green_fill,  font=Font(bold=True, color="006100")))

# Conditional formatting on Status column (G): 0=red 1=green
stat_range = f"G2:G{last_data_row}"
ws.conditional_formatting.add(stat_range, CellIsRule(operator="equal", formula=['0'], fill=red_fill))
ws.conditional_formatting.add(stat_range, CellIsRule(operator="equal", formula=['1'], fill=green_fill))

# Column widths
widths = {"A": 24, "B": 16, "C": 30, "D": 36, "E": 60, "F": 11, "G": 12, "H": 13}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ---------- Readiness Logic block ----------
ws.append([])
ws.append([])
logic_start = ws.max_row + 1

ws.cell(row=logic_start,     column=1, value="Readiness Logic (Manvitha 4/24 Teams)").font = Font(bold=True, size=12)
ws.cell(row=logic_start + 1, column=1, value="LOW Risk:").font = Font(bold=True, color="9C0006")
ws.cell(row=logic_start + 1, column=2, value="If ANY HIGH priority item Status = 0")
ws.cell(row=logic_start + 2, column=1, value="MEDIUM Risk:").font = Font(bold=True, color="9C5700")
ws.cell(row=logic_start + 2, column=2, value="If NO HIGH = 0 but ANY MEDIUM = 0")
ws.cell(row=logic_start + 3, column=1, value="HIGH (Ready):").font = Font(bold=True, color="006100")
ws.cell(row=logic_start + 3, column=2, value="If no HIGH or MEDIUM incomplete")

# ---------- Live computed scoring ----------
score_row = logic_start + 5
ws.cell(row=score_row,     column=1, value="Live Score").font = Font(bold=True, size=12)

ws.cell(row=score_row + 1, column=1, value="HIGH incomplete count:").font = Font(bold=True)
ws.cell(row=score_row + 1, column=2, value=f'=COUNTIFS(F2:F{last_data_row},"HIGH",G2:G{last_data_row},0)')

ws.cell(row=score_row + 2, column=1, value="MEDIUM incomplete count:").font = Font(bold=True)
ws.cell(row=score_row + 2, column=2, value=f'=COUNTIFS(F2:F{last_data_row},"MEDIUM",G2:G{last_data_row},0)')

ws.cell(row=score_row + 3, column=1, value="Computed Risk Tier:").font = Font(bold=True, size=12)
risk_formula = (
    f'=IF(COUNTIFS(F2:F{last_data_row},"HIGH",G2:G{last_data_row},0)>0,"LOW",'
    f'IF(COUNTIFS(F2:F{last_data_row},"MEDIUM",G2:G{last_data_row},0)>0,"MEDIUM",'
    f'"HIGH"))'
)
risk_cell = ws.cell(row=score_row + 3, column=2, value=risk_formula)
risk_cell.font = Font(bold=True, size=12)

# Conditional formatting on the computed risk cell
risk_range = f"B{score_row + 3}"
ws.conditional_formatting.add(risk_range, CellIsRule(operator="equal", formula=['"LOW"'],    fill=red_fill,    font=Font(bold=True, color="9C0006")))
ws.conditional_formatting.add(risk_range, CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=yellow_fill, font=Font(bold=True, color="9C5700")))
ws.conditional_formatting.add(risk_range, CellIsRule(operator="equal", formula=['"HIGH"'],   fill=green_fill,  font=Font(bold=True, color="006100")))

# Freeze header row
ws.freeze_panes = "A2"

# =================== SHEET 2: Business Rules ===================
ws2 = wb.create_sheet("Business Rules")
ws2.append(["Rule", "Description"])
business_rules = [
    ["Temporal Validation",   "Ignore LMS events before program_start_date (gates Course Login + Course Participation; EVENT_LAYER_SPEC sec 6.5)"],
    ["Census Rule",           "Remove student if today >= program_start_date + 1 day past census (Bryan corrected 4/27 @ 11:30 from prior 10-day figure)"],
    ["Negative Time Display", "If today > program_start_date show negative days post-start; zero-day handled separately"],
    ["FAFSA OR Alt-Funding",  "Checklist passes if EITHER fafsa_submission_flag=TRUE OR alt-funding intent recorded (v17.9 sec 11)"],
    ["WoW Dual Match",        "Match BOTH wow_login and wwow_login activity names (Bryan 4/27: 'leave it, cover both')"],
    ["Email/SMS Draft Blank", "NOT a bug - async AI generation populates 5-7s after page load (Jake 4/27 @ 30:00)"],
    ["Test Environment",      "QA = stable verification env once Jake permissions unblock (Kevin + Vishnu); supersedes 'verify in dev'"],
    ["Course Login Triage",   "If raw course_login=TRUE -> source bug -> Alpesh; if FALSE -> UI bug -> Jake (Jake 4/27 @ 29:19)"],
    ["NBA Grammarly Copy",    "Header: 'Set Up a Free Grammarly Account'; Sub: 'Guide student to Grammarly'; visible link styling"],
    ["Risk Tier (Manvitha 4/24)", "LOW=any HIGH incomplete; MEDIUM=no HIGH but any MEDIUM incomplete; HIGH=none incomplete"],
    ["Open: Accredited Piece","Bryan 4/27 @ 37:57 raised 'the accredited piece' - needs Jake sync 4/28"],
]
for r in business_rules:
    ws2.append(r)
for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 100
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
ws2.freeze_panes = "A2"

# =================== SHEET 3: Change Log ===================
ws3 = wb.create_sheet("Change Log")
ws3.append(["Date", "Author", "Change"])
log = [
    ["2026-04-24", "Manvitha (Teams)",            "Defined risk-tier logic: LOW/MEDIUM/HIGH from HIGH+MEDIUM checklist incompletes"],
    ["2026-04-27", "ChatGPT (Bryan v2 import)",   "Initial structure: Checklist Mapping + Business Rules"],
    ["2026-04-28", "ChatGPT (v4 export)",         "Added Priority, Status (0/1), Impacts Readiness columns + readiness logic block"],
    ["2026-04-28", "Nagendra (post Apr 27 call)", "Added LMS Login row (live 4/27, 5,423 rows)"],
    ["2026-04-28", "Nagendra",                    "WoW Login: explicit dual-match (Bryan)"],
    ["2026-04-28", "Nagendra",                    "Course Login: triage protocol + program-start gating (Jake)"],
    ["2026-04-28", "Nagendra",                    "Course Participation: confirmed by Jake live in call"],
    ["2026-04-28", "Nagendra",                    "FAFSA: OR-logic with alt-funding (v17.9 sec 11)"],
    ["2026-04-28", "Nagendra",                    "Census Rule: 10 days -> 1+ day past census (Bryan @ 11:30)"],
    ["2026-04-28", "Nagendra",                    "Added rules: Email/SMS async, Test env, NBA copy, accredited-piece open"],
    ["2026-04-28", "Nagendra (v5 final)",         "Color-coded Priority/Status; live formulas wire Status -> Computed Risk Tier"],
]
for r in log:
    ws3.append(r)
for cell in ws3[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    cell.border = border
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 90
for row in ws3.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
ws3.freeze_panes = "A2"

wb.save(DST)
print("Saved:", DST)

# Verify
wb2 = openpyxl.load_workbook(DST)
for s in wb2.worksheets:
    print(f"\n=== SHEET: {s.title} ===")
    for row in s.iter_rows(values_only=True):
        print(" | ".join(str(c) if c is not None else "" for c in row))
