from openpyxl import load_workbook
wb = load_workbook(r'context/daily/apr29/R2C_Checklist_apr29.xlsx')
for s in wb.worksheets:
    print('==SHEET==', s.title, 'dims=', s.dimensions)
    for r in s.iter_rows(values_only=True):
        print(r)
